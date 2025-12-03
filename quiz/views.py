# quiz/views.py

# ===== CÁC THƯ VIỆN CẦN THIẾT =====
# quiz/views.py

# ===== CÁC THƯ VIỆN CẦN THIẾT =====
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from django.db import transaction, models
from django.core.paginator import Paginator
from django.contrib import messages
import random
from openpyxl import load_workbook
from django.db.models import Count, Q
from django.core.exceptions import PermissionDenied
from .models import Question, Quiz, Answer, Subject
from support.models import SupportTicket
from users.models import User 


# ===== IMPORT TỪ CÁC FILE KHÁC TRONG DỰ ÁN =====
from .forms import QuestionForm, AnswerFormSet, QuizForm
from results.models import Result, StudentAnswer
from users.decorators import student_required, teacher_required
from results.models import Result, StudentAnswer
from results.models import PracticeResult
from django.db.models import Avg, F
import json
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.shortcuts import render, redirect
from users.decorators import student_required, teacher_required, admin_required  # THÊM admin_required
from django.utils import timezone

# ===========================================================================
# VIEWS CHUNG & PHÂN LUỒNG
# ===========================================================================

def landing_page(request):
    """View cho trang chủ công khai, ai cũng có thể xem được."""
    return render(request, 'pages/landing_page.html')


@login_required
def dashboard(request):
    user = request.user
    if user.role == 'TEACHER':
        return redirect('quiz:teacher_dashboard')
    
    elif user.role == 'STUDENT':
        now = timezone.now()
        # Lấy các đề thi công khai, đang diễn ra
        public_quizzes = Quiz.objects.filter(is_public=True, start_time__lte=now, end_time__gte=now)
        # Lấy các đề thi riêng tư mà học sinh đã được phép tham gia
        private_quizzes_joined = user.allowed_quizzes.filter(is_public=False, start_time__lte=now, end_time__gte=now)
        
        all_quizzes = (public_quizzes | private_quizzes_joined).distinct()
        
        taken_quiz_ids = Result.objects.filter(student=user).values_list('quiz__id', flat=True)
        context = {'quizzes': all_quizzes, 'taken_quiz_ids': list(taken_quiz_ids)}
        return render(request, 'pages/student_dashboard.html', context)
    
    else: # Admin
        return render(request, 'pages/admin_dashboard.html')


# ===========================================================================
# VIEWS CHO GIÁO VIÊN
# ===========================================================================

@login_required
@teacher_required
def teacher_dashboard(request):
    teacher = request.user
    total_questions = Question.objects.filter(created_by=teacher).count()
    total_quizzes = Quiz.objects.filter(created_by=teacher).count()
    recent_quizzes = Quiz.objects.filter(created_by=teacher).order_by('-created_at')[:5]
    total_attempts = Result.objects.filter(quiz__created_by=teacher).count()

    student_answers = StudentAnswer.objects.filter(result__quiz__created_by=teacher)
    total_answers_submitted = student_answers.count()
    total_correct_answers = student_answers.filter(selected_answer__is_correct=True).count()
    
    average_correct_rate = 0
    if total_answers_submitted > 0:
        average_correct_rate = round((total_correct_answers / total_answers_submitted) * 100)

    context = {
        'total_quizzes': total_quizzes, 
        'total_questions': total_questions, 
        'recent_quizzes': recent_quizzes, 
        'total_attempts': total_attempts, 
        'average_correct_rate': average_correct_rate
    }
    return render(request, 'pages/teacher_dashboard.html', context)


# ===========================================================================
# CHỨC NĂNG QUẢN LÝ CÂU HỎI (QUESTION CRUD)
# ===========================================================================

@login_required
@teacher_required
def question_list(request):
    questions_qs = Question.objects.filter(created_by=request.user)
    
    # Thêm filter
    subject_filter = request.GET.get('subject')
    difficulty_filter = request.GET.get('difficulty')
    question_type_filter = request.GET.get('question_type')
    search_query = request.GET.get('search', '')
    
    if subject_filter:
        questions_qs = questions_qs.filter(subject_id=subject_filter)
    if difficulty_filter:
        questions_qs = questions_qs.filter(difficulty=difficulty_filter)
    if question_type_filter:
        questions_qs = questions_qs.filter(question_type=question_type_filter)
    if search_query:
        questions_qs = questions_qs.filter(text__icontains=search_query)
    
    questions_qs = questions_qs.order_by('-created_at')
    
    # Thống kê
    total_questions = questions_qs.count()
    single_choice_count = questions_qs.filter(question_type='SINGLE_CHOICE').count()
    multiple_choice_count = questions_qs.filter(question_type='MULTIPLE_CHOICE').count()
    true_false_count = questions_qs.filter(question_type='TRUE_FALSE').count()
    short_answer_count = questions_qs.filter(question_type='SHORT_ANSWER').count()
    
    paginator = Paginator(questions_qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'subjects': Subject.objects.all(),
        'total_questions': total_questions,
        'single_choice_count': single_choice_count,
        'multiple_choice_count': multiple_choice_count,
        'true_false_count': true_false_count,
        'short_answer_count': short_answer_count,
        'search_query': search_query,
    }
    return render(request, 'quiz_management/question_list.html', context)


@login_required
@teacher_required
@transaction.atomic
def question_create(request):
    """View tạo câu hỏi mới - ĐÃ SỬA LỖI FORM"""
    if request.method == 'POST':
        form = QuestionForm(request.POST)
        formset = AnswerFormSet(request.POST)
        
        # Kiểm tra form chính trước
        if form.is_valid():
            question = form.save(commit=False)
            question.created_by = request.user
            question_type = form.cleaned_data['question_type']
            
            try:
                # Xử lý câu hỏi Đúng/Sai
                if question_type == Question.QuestionType.TRUE_FALSE:
                    question.save()
                    # Tạo tự động 2 đáp án Đúng/Sai
                    Answer.objects.create(question=question, text='Đúng', is_correct=True)
                    Answer.objects.create(question=question, text='Sai', is_correct=False)
                    messages.success(request, "Thêm câu hỏi Đúng/Sai thành công!")
                    return redirect('quiz:question_list')
                
                # Xử lý câu hỏi tự luận
                elif question_type == Question.QuestionType.SHORT_ANSWER:
                    question.save()
                    messages.success(request, "Thêm câu hỏi tự luận thành công!")
                    return redirect('quiz:question_list')
                
                # Xử lý câu hỏi có đáp án (Single/Multiple choice)
                else:
                    if formset.is_valid():
                        question.save()
                        formset.instance = question
                        formset.save()
                        messages.success(request, "Thêm câu hỏi thành công!")
                        return redirect('quiz:question_list')
                    else:
                        messages.error(request, "Vui lòng kiểm tra lại các đáp án.")
                        # GIỮ LẠI FORM VÀ FORMSET KHI CÓ LỖI
                        return render(request, 'quiz_management/question_form.html', {
                            'form': form, 
                            'formset': formset, 
                            'is_edit': False
                        })
                        
            except Exception as e:
                messages.error(request, f"Có lỗi xảy ra: {str(e)}")
                return render(request, 'quiz_management/question_form.html', {
                    'form': form, 
                    'formset': formset, 
                    'is_edit': False
                })
        else:
            messages.error(request, "Vui lòng kiểm tra lại thông tin câu hỏi.")
            # GIỮ LẠI FORM VÀ FORMSET KHI CÓ LỖI
            return render(request, 'quiz_management/question_form.html', {
                'form': form, 
                'formset': formset, 
                'is_edit': False
            })
    
    else:
        # GET request - khởi tạo form trống
        form = QuestionForm()
        formset = AnswerFormSet()
    
    return render(request, 'quiz_management/question_form.html', {
        'form': form, 
        'formset': formset, 
        'is_edit': False
    })

# === API TẠO CÂU HỎI NHANH (AJAX) ===
@login_required
@teacher_required
@require_POST
def api_quick_create_question(request):
    try:
        data = json.loads(request.body)
        
        # 1. Lấy dữ liệu từ Modal
        text = data.get('text')
        q_type = data.get('question_type')
        subject_id = data.get('subject_id')
        difficulty = data.get('difficulty', 'MEDIUM')
        
        # Dữ liệu cho câu tự luận / trắc nghiệm
        correct_answer_text = data.get('correct_answer_text', '') # Cho tự luận
        answers_data = data.get('answers', []) # Cho trắc nghiệm [{'text': 'A', 'is_correct': True}, ...]

        # 2. Tạo câu hỏi
        subject = Subject.objects.get(id=subject_id)
        question = Question.objects.create(
            text=text,
            subject=subject,
            question_type=q_type,
            difficulty=difficulty,
            created_by=request.user,
            correct_answer_text=correct_answer_text if q_type == 'SHORT_ANSWER' else None
        )

        # 3. Tạo đáp án (Nếu là trắc nghiệm)
        if q_type in ['SINGLE_CHOICE', 'MULTIPLE_CHOICE']:
            for ans in answers_data:
                if ans.get('text'): # Chỉ lưu đáp án có nội dung
                    Answer.objects.create(
                        question=question,
                        text=ans.get('text'),
                        is_correct=ans.get('is_correct', False)
                    )
        
        # 4. Trả về dữ liệu để hiển thị ngay lập tức
        return JsonResponse({
            'status': 'success',
            'question': {
                'id': question.id,
                'text': question.text,
                'difficulty': question.get_difficulty_display(),
                'type': question.get_question_type_display(),
                'subject': question.subject.name
            }
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
@login_required
@teacher_required
@transaction.atomic
def question_edit(request, pk):
    """View chỉnh sửa câu hỏi"""
    question = get_object_or_404(Question, pk=pk, created_by=request.user)
    
    if request.method == 'POST':
        form = QuestionForm(request.POST, instance=question)
        
        # Chỉ tạo formset cho các câu hỏi có đáp án
        if question.question_type not in [Question.QuestionType.TRUE_FALSE, Question.QuestionType.SHORT_ANSWER]:
            formset = AnswerFormSet(request.POST, instance=question)
        else:
            formset = None
        
        if form.is_valid():
            try:
                # Xử lý câu hỏi Đúng/Sai - không cho phép chỉnh sửa đáp án
                if question.question_type == Question.QuestionType.TRUE_FALSE:
                    form.save()
                    messages.success(request, 'Cập nhật câu hỏi Đúng/Sai thành công!')
                    return redirect('quiz:question_list')
                
                # Xử lý câu hỏi tự luận
                elif question.question_type == Question.QuestionType.SHORT_ANSWER:
                    form.save()
                    messages.success(request, 'Cập nhật câu hỏi tự luận thành công!')
                    return redirect('quiz:question_list')
                
                # Xử lý các câu hỏi có đáp án khác
                else:
                    if formset and formset.is_valid():
                        form.save()
                        formset.save()
                        messages.success(request, 'Cập nhật câu hỏi thành công!')
                        return redirect('quiz:question_list')
                    else:
                        messages.error(request, "Vui lòng kiểm tra lại các đáp án.")
                        return render(request, 'quiz_management/question_form.html', {
                            'form': form, 
                            'formset': formset, 
                            'is_edit': True,
                            'question': question
                        })
            except Exception as e:
                messages.error(request, f"Có lỗi xảy ra: {str(e)}")
        else:
            messages.error(request, "Vui lòng kiểm tra lại thông tin câu hỏi.")
    
    else:
        form = QuestionForm(instance=question)
        
        # Chỉ tạo formset cho các câu hỏi có đáp án
        if question.question_type not in [Question.QuestionType.TRUE_FALSE, Question.QuestionType.SHORT_ANSWER]:
            formset = AnswerFormSet(instance=question)
        else:
            formset = None
    
    context = {
        'form': form, 
        'formset': formset, 
        'is_edit': True,
        'question': question
    }
    return render(request, 'quiz_management/question_form.html', context)


@login_required
@teacher_required
def question_delete(request, pk):
    """View xóa câu hỏi"""
    question = get_object_or_404(Question, pk=pk, created_by=request.user)
    if request.method == 'POST':
        question.delete()
        messages.success(request, 'Xóa câu hỏi thành công!')
        return redirect('quiz:question_list')
    return render(request, 'quiz_management/question_confirm_delete.html', {'question': question})


@login_required
@teacher_required
def question_import_excel(request):
    """View import câu hỏi từ file Excel"""
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file or not excel_file.name.endswith('.xlsx'):
            messages.error(request, "Vui lòng upload một file Excel (.xlsx).")
            return redirect('quiz:question_import')
        
        try:
            workbook = load_workbook(excel_file)
            sheet = workbook.active
            count = 0
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Kiểm tra dòng có đủ dữ liệu không
                if not all(row[i] for i in [0, 1, 2, 3, 4, 7]):
                    continue
                    
                subject_name, q_text, difficulty, ans1, ans2, ans3, ans4, correct_idx_str = row[:8]
                subject, _ = Subject.objects.get_or_create(name=str(subject_name).strip())
                
                difficulty_val = str(difficulty).strip().upper()
                if difficulty_val not in Question.Difficulty.values:
                    continue
                
                # Tạo câu hỏi với loại mặc định là SINGLE_CHOICE
                question = Question.objects.create(
                    subject=subject, 
                    text=str(q_text).strip(), 
                    difficulty=difficulty_val, 
                    question_type=Question.QuestionType.SINGLE_CHOICE,
                    created_by=request.user
                )
                
                answers = [ans1, ans2, ans3, ans4]
                correct_idx = int(correct_idx_str)
                
                for i, ans_text in enumerate(answers, 1):
                    if ans_text:
                        Answer.objects.create(
                            question=question, 
                            text=str(ans_text).strip(), 
                            is_correct=(i == correct_idx)
                        )
                count += 1
                
            messages.success(request, f'Import thành công {count} câu hỏi.')
            return redirect('quiz:question_list')
            
        except Exception as e:
            messages.error(request, f"Lỗi xử lý file Excel: {e}")
            return redirect('quiz:question_import')
    
    return render(request, 'quiz_management/question_import.html')
# ===========================================================================
# CHỨC NĂNG QUẢN LÝ ĐỀ THI (QUIZ CRUD)
# ===========================================================================

@login_required
@teacher_required
def quiz_list(request):
    """View danh sách đề thi"""
    quizzes = Quiz.objects.filter(created_by=request.user).order_by('-created_at')
    context = {
        'quizzes': quizzes,
        'now': timezone.now()
    }
    return render(request, 'quiz_management/quiz_list.html', context)


@login_required
@teacher_required
def quiz_create(request):
    """View tạo đề thi mới"""
    if request.method == "POST":
        form = QuizForm(request.POST, user=request.user)
        if form.is_valid():
            try:
                quiz = form.save(commit=False)
                quiz.created_by = request.user
                quiz.save()
                form.save_m2m()  # Quan trọng để lưu ManyToManyField
                messages.success(request, "Tạo đề thi thành công!")
                return redirect('quiz:quiz_list')
            except Exception as e:
                messages.error(request, f"Có lỗi xảy ra: {str(e)}")
        else:
            messages.error(request, "Vui lòng kiểm tra lại thông tin đề thi.")
    else:
        form = QuizForm(user=request.user)
    
    return render(request, 'quiz_management/quiz_form.html', {'form': form, 'is_edit': False})


@login_required
@teacher_required
def quiz_edit(request, pk):
    """View chỉnh sửa đề thi"""
    quiz = get_object_or_404(Quiz, pk=pk, created_by=request.user)
    
    if request.method == 'POST':
        form = QuizForm(request.POST, instance=quiz, user=request.user)
        if form.is_valid():
            try:
                quiz = form.save(commit=False)
                quiz.save()
                form.save_m2m()
                messages.success(request, 'Cập nhật đề thi thành công!')
                return redirect('quiz:quiz_list')
            except Exception as e:
                messages.error(request, f"Có lỗi xảy ra: {str(e)}")
        else:
            messages.error(request, "Vui lòng kiểm tra lại thông tin đề thi.")
    else:
        form = QuizForm(instance=quiz, user=request.user)
    
    return render(request, 'quiz_management/quiz_form.html', {
        'form': form, 
        'is_edit': True, 
        'quiz': quiz
    })


@login_required
@teacher_required
def quiz_delete(request, pk):
    """View xóa đề thi"""
    quiz = get_object_or_404(Quiz, pk=pk, created_by=request.user)
    if request.method == 'POST':
        quiz.delete()
        messages.success(request, 'Xóa đề thi thành công!')
        return redirect('quiz:quiz_list')
    return render(request, 'quiz_management/quiz_confirm_delete.html', {'quiz': quiz})


@login_required
@teacher_required
def quiz_results(request, pk):
    """View kết quả đề thi"""
    quiz = get_object_or_404(Quiz, pk=pk, created_by=request.user)
    results = Result.objects.filter(quiz=quiz).order_by('-score')
    return render(request, 'quiz_management/quiz_results.html', {
        'quiz': quiz, 
        'results': results
    })


# ===========================================================================
# VIEWS CHO HỌC SINH
# ===========================================================================

@login_required
@student_required
def join_with_code(request):
    """View tham gia đề thi bằng mã code"""
    if request.method == 'POST':
        code = request.POST.get('access_code', '').strip().upper()
        if not code:
            messages.error(request, "Vui lòng nhập mã tham gia.")
            return redirect('dashboard')
        
        try:
            quiz = Quiz.objects.get(access_code=code)
            quiz.allowed_students.add(request.user)
            messages.success(request, f"Bạn đã tham gia thành công vào đề thi '{quiz.title}'.")
        except Quiz.DoesNotExist:
            messages.error(request, "Mã tham gia không hợp lệ hoặc không tồn tại.")
    
    return redirect('dashboard')


@login_required
@student_required
def take_quiz(request, pk):
    """View làm bài thi"""
    quiz = get_object_or_404(Quiz, pk=pk)
    now = timezone.now()
    
    # Kiểm tra quyền truy cập
    if not quiz.is_public and request.user not in quiz.allowed_students.all():
        messages.error(request, "Đây là kỳ thi riêng tư. Bạn cần nhập mã tham gia trước.")
        return redirect('dashboard')
    
    # KIỂM TRA SỐ LẦN THI - CHỈ ĐỂ 1 PHẦN NÀY
    if not quiz.allow_multiple_attempts:
        # Nếu chỉ cho phép thi 1 lần, kiểm tra xem học sinh đã thi chưa
        existing_result = Result.objects.filter(quiz=quiz, student=request.user).first()
        if existing_result:
            messages.warning(request, "Bạn đã hoàn thành đề thi này rồi!")
            return redirect('quiz:view_result', pk=existing_result.id)
    else:
        # Nếu cho phép thi nhiều lần, vẫn cho phép thi lại
        pass
    
    # Kiểm tra thời gian
    if now > quiz.end_time:
        messages.error(request, "Bài thi này đã kết thúc.")
        return redirect('dashboard')
    if now < quiz.start_time:
        messages.info(request, "Bài thi này chưa đến giờ bắt đầu.")
        return redirect('dashboard')
    
    # Lấy và xáo trộn câu hỏi
    questions = list(quiz.questions.all())
    random.shuffle(questions)
    
    shuffled_questions = []
    for q in questions:
        if q.question_type in [Question.QuestionType.SINGLE_CHOICE, Question.QuestionType.MULTIPLE_CHOICE, Question.QuestionType.TRUE_FALSE]:
            # Xáo trộn đáp án cho câu hỏi có lựa chọn
            answers = list(q.answers.all())
            random.shuffle(answers)
            shuffled_questions.append({
                'question': q, 
                'answers': answers,
                'question_type': q.question_type
            })
        else:
            # Câu hỏi tự luận - không có đáp án để xáo trộn
            shuffled_questions.append({
                'question': q,
                'answers': [],
                'question_type': q.question_type
            })

    context = {
        'quiz': quiz, 
        'shuffled_questions': shuffled_questions,
        'now': now,
        'end_time': quiz.end_time
    }
    return render(request, 'quiz_taking/take_quiz.html', context)

@login_required
@student_required
@transaction.atomic
def submit_quiz(request, pk):
    """View nộp bài thi - ĐÃ SỬA LỖI XỬ LÝ CÁC LOẠI CÂU HỎI"""
    if request.method != 'POST': 
        raise PermissionDenied
    
    quiz = get_object_or_404(Quiz, pk=pk)
    now = timezone.now()
    
    if now > quiz.end_time: 
        messages.error(request, "Đã hết thời gian làm bài, không thể nộp.")
        return redirect('dashboard')
    
    if Result.objects.filter(student=request.user, quiz=quiz).exists(): 
        messages.warning(request, "Bạn đã nộp bài thi này rồi.")
        return redirect('dashboard')

    questions = quiz.questions.all()
    correct_answers_count = 0
    total_questions = questions.count()
    result = Result.objects.create(student=request.user, quiz=quiz, score=0)
    
    for question in questions:
        if question.question_type == Question.QuestionType.SINGLE_CHOICE:
            # Xử lý câu hỏi một lựa chọn
            selected_answer_id = request.POST.get(f'question_{question.id}')
            if selected_answer_id:
                try:
                    selected_answer = Answer.objects.get(pk=selected_answer_id)
                    StudentAnswer.objects.create(
                        result=result, 
                        question=question, 
                        selected_answer=selected_answer
                    )
                    if selected_answer.is_correct: 
                        correct_answers_count += 1
                except Answer.DoesNotExist: 
                    pass

        elif question.question_type == Question.QuestionType.MULTIPLE_CHOICE:
            # Xử lý câu hỏi nhiều lựa chọn
            selected_answer_ids = request.POST.getlist(f'question_{question.id}')
            if selected_answer_ids:
                # Lấy tất cả đáp án đúng của câu hỏi
                correct_answers = set(question.answers.filter(is_correct=True).values_list('id', flat=True))
                selected_answers = set(int(id) for id in selected_answer_ids)
                
                # Tạo StudentAnswer cho mỗi đáp án được chọn
                for answer_id in selected_answers:
                    try:
                        selected_answer = Answer.objects.get(pk=answer_id)
                        StudentAnswer.objects.create(
                            result=result, 
                            question=question, 
                            selected_answer=selected_answer
                        )
                    except Answer.DoesNotExist:
                        pass
                
                # Kiểm tra nếu tất cả đáp án đúng được chọn và không có đáp án sai nào được chọn
                if selected_answers == correct_answers:
                    correct_answers_count += 1

        elif question.question_type == Question.QuestionType.TRUE_FALSE:
            # Xử lý câu hỏi Đúng/Sai - SỬA LẠI PHẦN NÀY
            selected_value = request.POST.get(f'question_{question.id}')
            if selected_value:
                # Tìm đáp án đúng
                correct_answer = question.answers.filter(is_correct=True).first()
                if correct_answer:
                    # Tìm đáp án tương ứng với lựa chọn của học sinh
                    try:
                        selected_answer = question.answers.get(
                            text='Đúng' if selected_value.lower() == 'true' else 'Sai'
                        )
                        StudentAnswer.objects.create(
                            result=result, 
                            question=question, 
                            selected_answer=selected_answer
                        )
                        # Kiểm tra nếu đáp án học sinh chọn là đáp án đúng
                        if selected_answer.is_correct:
                            correct_answers_count += 1
                    except Answer.DoesNotExist:
                        pass

        elif question.question_type == Question.QuestionType.SHORT_ANSWER:
            # Xử lý câu hỏi tự luận ngắn
            student_answer_text = request.POST.get(f'short_answer_{question.id}', '').strip()
            if student_answer_text:
                StudentAnswer.objects.create(
                    result=result, 
                    question=question, 
                    selected_answer=None,
                    custom_answer=student_answer_text  # Lưu câu trả lời tự luận
                )
            # Câu tự luận không tính điểm tự động
    
    # Tính điểm (chỉ tính các câu hỏi có thể chấm tự động)
    scorable_questions = questions.exclude(question_type=Question.QuestionType.SHORT_ANSWER)
    total_scorable = scorable_questions.count()
    
    if total_scorable > 0:
        score = (correct_answers_count / total_scorable) * 100
    else:
        score = 0
        
    result.score = round(score, 2)
    result.save()
    
    messages.success(request, "Nộp bài thi thành công!")
    return redirect('quiz:view_result', pk=result.pk)

@login_required
@student_required
def view_result(request, pk):
    """View xem kết quả bài thi - ĐÃ SỬA LỖI HIỂN THỊ"""
    result = get_object_or_404(Result, pk=pk, student=request.user)
    detailed_answers = []
    student_answers_dict = {sa.question.id: sa for sa in result.student_answers.all()}
    
    for question in result.quiz.questions.order_by('id'):
        student_answer_obj = student_answers_dict.get(question.id)
        
        if question.question_type == Question.QuestionType.SHORT_ANSWER:
            # Xử lý câu hỏi tự luận
            detailed_answers.append({
                'question': question,
                'question_type': question.question_type,
                'correct_answer_text': question.correct_answer_text,
                'student_answer_text': student_answer_obj.custom_answer if student_answer_obj else '',
                'is_short_answer': True,
                'is_multiple_choice': False,
                'explanation': question.explanation
            })
        elif question.question_type == Question.QuestionType.MULTIPLE_CHOICE:
            # Xử lý câu hỏi nhiều lựa chọn
            correct_answers = question.answers.filter(is_correct=True)
            student_selected_answers = []
            if student_answer_obj:
                # Lấy tất cả câu trả lời của học sinh cho câu hỏi này
                student_selected_answers = list(result.student_answers.filter(
                    question=question, 
                    selected_answer__isnull=False
                ).values_list('selected_answer_id', flat=True))
            
            # Kiểm tra đúng/sai cho multiple choice
            correct_answer_ids = set(correct_answers.values_list('id', flat=True))
            student_answer_ids = set(student_selected_answers)
            is_correct = (correct_answer_ids == student_answer_ids)
            
            detailed_answers.append({
                'question': question,
                'question_type': question.question_type,
                'all_answers': question.answers.all(),
                'correct_answers': correct_answers,
                'student_selected_answers': student_selected_answers,
                'is_correct': is_correct,
                'is_short_answer': False,
                'is_multiple_choice': True,
                'explanation': question.explanation
                
            })
        else:
            # Xử lý câu hỏi một lựa chọn và Đúng/Sai
            correct_answer = question.answers.filter(is_correct=True).first()
            student_answer = student_answer_obj.selected_answer if student_answer_obj else None
            is_correct = student_answer.is_correct if student_answer else False
            
            detailed_answers.append({
                'question': question,
                'question_type': question.question_type,
                'all_answers': question.answers.all(),
                'correct_answer': correct_answer,
                'student_answer': student_answer,
                'is_correct': is_correct,
                'is_short_answer': False,
                'is_multiple_choice': False,
                'explanation': question.explanation
            })
    context = {
        'result': result, 
        'detailed_answers': detailed_answers
    }
    return render(request, 'quiz_taking/view_result.html', context)

@login_required
@student_required
def test_history(request):
    """View lịch sử làm bài"""
    results = Result.objects.filter(student=request.user).order_by('-completed_at')
    context = {'results': results}
    return render(request, 'quiz_taking/test_history.html', context)


@login_required
def practice_selection(request):
    """Trang chọn đề để luyện tập"""
    # Lấy tất cả đề thi cho phép luyện tập nhiều lần
    available_quizzes = Quiz.objects.filter(
        allow_multiple_attempts=True,
        is_public=True
    ).distinct()
    
    # Thêm thông tin thống kê cho mỗi quiz
    for quiz in available_quizzes:
        # Số lần luyện tập
        quiz.attempt_count = PracticeResult.objects.filter(
            quiz=quiz, 
            student=request.user
        ).count()
        
        # Điểm cao nhất
        best_result = PracticeResult.objects.filter(
            quiz=quiz, 
            student=request.user
        ).order_by('-score').first()
        quiz.best_score = best_result.score if best_result else None
    
    # Tính toán thống kê tổng
    total_attempts = PracticeResult.objects.filter(student=request.user).count()
    user_practice_results = PracticeResult.objects.filter(student=request.user)
    
    if user_practice_results.exists():
        average_score = user_practice_results.aggregate(Avg('score'))['score__avg']
        average_score = round(average_score, 1) if average_score else 0
    else:
        average_score = 0
    
    # Tính % cải thiện
    improvement = 0
    recent_results = user_practice_results.order_by('-completed_at')[:2]
    if len(recent_results) == 2:
        old_score = recent_results[1].score
        new_score = recent_results[0].score
        if old_score > 0:
            improvement = round(((new_score - old_score) / old_score) * 100, 1)
    
    context = {
        'available_quizzes': available_quizzes,
        'total_attempts': total_attempts,
        'average_score': average_score,
        'total_quizzes': available_quizzes.count(),
        'improvement': improvement,
    }
    return render(request, 'practice_mode/practice_selection.html', context)

@login_required
def submit_practice_quiz(request, pk):
    """Xử lý nộp bài luyện tập"""
    if request.method == 'POST':
        quiz = get_object_or_404(Quiz, pk=pk)
        
        if not quiz.allow_multiple_attempts:
            messages.error(request, "Đề thi này không cho phép luyện tập!")
            return redirect('quiz:practice_selection')
        
        # Tính điểm và chi tiết câu trả lời
        score = calculate_practice_score(request, quiz)
        detailed_answers = get_practice_detailed_answers(request, quiz)
        correct_count = len([a for a in detailed_answers if a['is_correct']])
        
        # Lưu kết quả luyện tập
        practice_result = PracticeResult.objects.create(
            student=request.user,
            quiz=quiz,
            score=score,
            total_questions=quiz.questions.count(),
            correct_answers=correct_count
        )
        
        context = {
            'result': {
                'quiz': quiz,
                'score': score,
                'detailed_answers': detailed_answers,
                'total_questions': quiz.questions.count(),
                'correct_answers': correct_count,
                'improvement': calculate_improvement(request.user, quiz, score),
            }
        }
        
        messages.success(request, f"Hoàn thành luyện tập! Điểm của bạn: {score}/100")
        return render(request, 'practice_mode/practice_result.html', context)
    
    return redirect('quiz:practice_selection')

def calculate_improvement(user, quiz, current_score):
    """Tính % cải thiện so với lần trước"""
    previous_results = PracticeResult.objects.filter(
        student=user, 
        quiz=quiz
    ).order_by('-completed_at')
    
    if previous_results.count() >= 2:
        previous_score = previous_results[1].score  # Lần trước đó
        if previous_score > 0:
            return round(((current_score - previous_score) / previous_score) * 100, 1)
    return 0
@login_required
def practice_quiz(request, pk):
    """Làm bài luyện tập"""
    quiz = get_object_or_404(Quiz, pk=pk)
    
    # Kiểm tra xem đề có cho phép luyện tập không
    if not quiz.allow_multiple_attempts:
        messages.error(request, "Đề thi này không cho phép luyện tập!")
        return redirect('quiz:practice_selection')
    
    # Lấy và xáo trộn câu hỏi
    questions = list(quiz.questions.all())
    random.shuffle(questions)
    
    shuffled_questions = []
    for q in questions:
        if q.question_type in [Question.QuestionType.SINGLE_CHOICE, Question.QuestionType.MULTIPLE_CHOICE, Question.QuestionType.TRUE_FALSE]:
            # Xáo trộn đáp án cho câu hỏi có lựa chọn
            answers = list(q.answers.all())
            random.shuffle(answers)
            shuffled_questions.append({
                'question': q, 
                'answers': answers,
                'question_type': q.question_type
            })
        else:
            # Câu hỏi tự luận - không có đáp án để xáo trộn
            shuffled_questions.append({
                'question': q,
                'answers': [],
                'question_type': q.question_type
            })

    context = {
        'quiz': quiz,
        'shuffled_questions': shuffled_questions,
        'questions': questions,  # Cho progress bar
        'is_practice': True,    # Đánh dấu là chế độ luyện tập
    }
    return render(request, 'practice_mode/practice_session.html', context)

@login_required
def practice_history(request):
    """Lịch sử luyện tập"""
    practice_results = PracticeResult.objects.filter(student=request.user).order_by('-completed_at')
    
    context = {
        'practice_results': practice_results,
    }
    return render(request, 'practice_mode/practice_history.html', context)

@login_required
def practice_by_subject(request, subject_id):
    """Luyện tập theo chủ đề môn học"""
    subject = get_object_or_404(Subject, pk=subject_id)
    available_quizzes = Quiz.objects.filter(
        subject=subject,
        allow_multiple_attempts=True,
        is_public=True
    )
    
    context = {
        'subject': subject,
        'available_quizzes': available_quizzes,
    }
    return render(request, 'practice_mode/practice_by_subject.html', context)

@login_required
def practice_random(request):
    """Luyện tập ngẫu nhiên"""
    # Lấy một đề thi ngẫu nhiên cho phép luyện tập
    available_quizzes = Quiz.objects.filter(
        allow_multiple_attempts=True,
        is_public=True
    )
    
    if available_quizzes.exists():
        random_quiz = random.choice(available_quizzes)
        return redirect('quiz:practice_quiz', pk=random_quiz.id)
    else:
        messages.info(request, "Hiện không có đề thi nào để luyện tập.")
        return redirect('quiz:practice_selection')

# ===========================================================================
# HÀM HỖ TRỢ TÍNH ĐIỂM CHO LUYỆN TẬP
# ===========================================================================

def calculate_practice_score(request, quiz):
    """Tính điểm cho bài luyện tập"""
    score = 0
    total_questions = quiz.questions.count()
    
    if total_questions == 0:
        return 0
    
    for question in quiz.questions.all():
        if is_answer_correct(request, question):
            score += 1
    
    return round((score / total_questions) * 100, 2)

def get_practice_detailed_answers(request, quiz):
    """Lấy chi tiết câu trả lời cho kết quả luyện tập"""
    detailed_answers = []
    
    for question in quiz.questions.all():
        student_answer = get_student_answer_from_request(request, question)
        is_correct = is_answer_correct(request, question)
        
        detailed_answers.append({
            'question': question,
            'student_answer': student_answer,
            'is_correct': is_correct,
            'is_short_answer': question.question_type == Question.QuestionType.SHORT_ANSWER,
            'student_answer_text': student_answer.text if student_answer else request.POST.get(f'question_{question.id}', ''),
            'correct_answer_text': question.correct_answer_text if question.question_type == Question.QuestionType.SHORT_ANSWER else '',
            'correct_answer': get_correct_answer(question),
            'all_answers': question.answers.all() if question.question_type != Question.QuestionType.SHORT_ANSWER else [],
        })
    
    return detailed_answers

def is_answer_correct(request, question):
    """Kiểm tra câu trả lời có đúng không"""
    if question.question_type == Question.QuestionType.SHORT_ANSWER:
        # Với câu tự luận, luôn tính là đúng trong luyện tập (hoặc có thể bỏ qua)
        return True
    elif question.question_type == Question.QuestionType.SINGLE_CHOICE:
        selected_answer_id = request.POST.get(f'question_{question.id}')
        if selected_answer_id:
            try:
                selected_answer = Answer.objects.get(id=selected_answer_id)
                return selected_answer.is_correct
            except Answer.DoesNotExist:
                return False
        return False
    elif question.question_type == Question.QuestionType.MULTIPLE_CHOICE:
        selected_ids = request.POST.getlist(f'question_{question.id}')
        correct_answers = question.answers.filter(is_correct=True)
        selected_correct = Answer.objects.filter(id__in=selected_ids, is_correct=True)
        return selected_correct.count() == correct_answers.count() and len(selected_ids) == correct_answers.count()
    elif question.question_type == Question.QuestionType.TRUE_FALSE:
        selected_answer_id = request.POST.get(f'question_{question.id}')
        if selected_answer_id:
            try:
                selected_answer = Answer.objects.get(id=selected_answer_id)
                return selected_answer.is_correct
            except Answer.DoesNotExist:
                return False
        return False
    return False

def get_student_answer_from_request(request, question):
    """Lấy câu trả lời của học sinh từ request"""
    if question.question_type == Question.QuestionType.SINGLE_CHOICE:
        answer_id = request.POST.get(f'question_{question.id}')
        if answer_id:
            return Answer.objects.filter(id=answer_id).first()
    elif question.question_type == Question.QuestionType.MULTIPLE_CHOICE:
        # Trả về answer đầu tiên được chọn (cho đơn giản)
        answer_ids = request.POST.getlist(f'question_{question.id}')
        if answer_ids:
            return Answer.objects.filter(id=answer_ids[0]).first()
    elif question.question_type == Question.QuestionType.TRUE_FALSE:
        answer_id = request.POST.get(f'question_{question.id}')
        if answer_id:
            return Answer.objects.filter(id=answer_id).first()
    return None

def get_correct_answer(question):
    """Lấy đáp án đúng của câu hỏi"""
    if question.question_type == Question.QuestionType.SINGLE_CHOICE:
        return question.answers.filter(is_correct=True).first()
    elif question.question_type == Question.QuestionType.MULTIPLE_CHOICE:
        return question.answers.filter(is_correct=True).first()  # Trả về 1 đáp án đúng
    elif question.question_type == Question.QuestionType.TRUE_FALSE:
        return question.answers.filter(is_correct=True).first()
    return None
@teacher_required
def grading_dashboard(request):
    """Dashboard chấm điểm cho giáo viên"""
    # Lấy các bài thi có câu hỏi tự luận chưa chấm
    pending_results = Result.objects.filter(
        quiz__questions__question_type='SHORT_ANSWER',
        is_graded=False
    ).distinct()
    
    # Lấy các bài thi đã chấm
    graded_results = Result.objects.filter(
        quiz__questions__question_type='SHORT_ANSWER',
        is_graded=True
    ).distinct()
    
    # Thêm số câu tự luận cho mỗi result
    # for result in pending_results:
    #     result.short_answer_count = result.student_answers.filter(
    #         question__question_type='SHORT_ANSWER'
    #     ).count()
    
    # for result in graded_results:
    #     result.short_answer_count = result.student_answers.filter(
    #         question__question_type='SHORT_ANSWER'
    #     ).count()
    
    context = {
        'pending_results': pending_results,
        'graded_results': graded_results,
    }
    return render(request, 'quiz_grading/grading_dashboard.html', context)

@teacher_required
def grade_short_answer(request, result_id):
    """View để giáo viên chấm điểm câu hỏi tự luận"""
    result = get_object_or_404(Result, id=result_id)
    
    if request.method == 'POST':
        total_points = 0
        short_answer_count = 0
        
        # Xử lý điểm cho từng câu hỏi tự luận
        for student_answer in result.student_answers.filter(question__question_type='SHORT_ANSWER'):
            points_field = f"points_{student_answer.id}"
            comment_field = f"comment_{student_answer.id}"
            
            if points_field in request.POST:
                try:
                    points = float(request.POST[points_field])
                    comment = request.POST.get(comment_field, '')
                    
                    # Cập nhật điểm và nhận xét
                    student_answer.points_earned = points
                    student_answer.teacher_comment = comment
                    student_answer.save()
                    
                    total_points += points
                    short_answer_count += 1
                    
                except ValueError:
                    messages.error(request, f"Điểm cho câu {student_answer.id} không hợp lệ")
        
        # Cập nhật kết quả
        result.short_answer_score = total_points
        result.is_graded = True
        result.teacher_feedback = request.POST.get('overall_feedback', '')
        result.save()
        
        messages.success(request, f'Đã chấm điểm {short_answer_count} câu tự luận! Tổng điểm tự luận: {total_points}')
        return redirect('quiz:grading_dashboard')
    
    # Lấy tất cả câu trả lời tự luận
    short_answers = result.student_answers.filter(question__question_type='SHORT_ANSWER')
    
    context = {
        'result': result,
        'short_answers': short_answers,
    }
    return render(request, 'quiz_grading/grade_answers.html', context)

# ===========================================================================
# VIEWS CÀI ĐẶT VÀ HỖ TRỢ
# ===========================================================================

@login_required
def settings_page(request):
    """Trang cài đặt chung với các tab"""
    user = request.user
    active_tab = request.GET.get('tab', 'profile')  # Mặc định là tab profile
    
    # Lấy thống kê thật cho hiển thị
    if user.role == 'TEACHER':
        total_questions = Question.objects.filter(created_by=user).count()
        total_quizzes = Quiz.objects.filter(created_by=user).count()
        total_attempts = Result.objects.filter(quiz__created_by=user).count()
        avg_score = Result.objects.filter(quiz__created_by=user).aggregate(Avg('score'))['score__avg'] or 0
        
        context = {
            'user': user,
            'active_tab': active_tab,
            'role': user.role,
            'total_questions': total_questions,
            'total_quizzes': total_quizzes,
            'total_attempts': total_attempts,
            'average_correct_rate': round(avg_score, 1) if avg_score else 0,
            'current_date': timezone.now(),  # Thêm thời gian hiện tại
        }
    else:  # STUDENT
        total_quizzes_taken = Result.objects.filter(student=user).count()
        total_practice = PracticeResult.objects.filter(student=user).count()
        avg_result = Result.objects.filter(student=user).aggregate(Avg('score'))['score__avg'] or 0
        
        context = {
            'user': user,
            'active_tab': active_tab,
            'role': user.role,
            'total_quizzes_taken': total_quizzes_taken,
            'total_practice': total_practice,
            'avg_score': round(avg_result, 1) if avg_result else 0,
            'current_date': timezone.now(),  # Thêm thời gian hiện tại
        }
    
    return render(request, 'settings/base_settings.html', context)
@login_required
def update_profile(request):
    """Cập nhật thông tin cá nhân"""
    if request.method == 'POST':
        try:
            user = request.user
            # Lấy dữ liệu từ form
            user.first_name = request.POST.get('first_name', '').strip()
            user.last_name = request.POST.get('last_name', '').strip()
            user.email = request.POST.get('email', '').strip()
            
            # Xử lý avatar nếu có
            if 'avatar' in request.FILES:
                user.avatar = request.FILES['avatar']
            
            # Lưu thông tin bổ sung theo role (nếu model có các trường này)
            if hasattr(user, 'student_id'):
                user.student_id = request.POST.get('student_id', '').strip()
            if hasattr(user, 'class_name'):
                user.class_name = request.POST.get('class_name', '').strip()
            if hasattr(user, 'school'):
                user.school = request.POST.get('school', '').strip()
            if hasattr(user, 'teacher_id'):
                user.teacher_id = request.POST.get('teacher_id', '').strip()
            if hasattr(user, 'department'):
                user.department = request.POST.get('department', '').strip()
            
            user.save()
            messages.success(request, "Cập nhật thông tin thành công!")
            
        except Exception as e:
            messages.error(request, f"Có lỗi xảy ra: {str(e)}")
    
    return redirect('quiz:settings')

@login_required
def change_password(request):
    """Thay đổi mật khẩu"""
    if request.method == 'POST':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        user = request.user
        
        # Kiểm tra mật khẩu hiện tại
        if not user.check_password(current_password):
            messages.error(request, "Mật khẩu hiện tại không đúng!")
            return redirect('quiz:settings')
        
        # Kiểm tra mật khẩu mới
        if new_password != confirm_password:
            messages.error(request, "Mật khẩu xác nhận không khớp!")
            return redirect('quiz:settings')
        
        if len(new_password) < 6:
            messages.error(request, "Mật khẩu phải có ít nhất 6 ký tự!")
            return redirect('quiz:settings')
        
        # Cập nhật mật khẩu
        user.set_password(new_password)
        user.save()
        
        # Cập nhật lại session
        from django.contrib.auth import update_session_auth_hash
        update_session_auth_hash(request, user)
        
        messages.success(request, "Đổi mật khẩu thành công!")
    
    return redirect('quiz:settings')

# ===========================================================================
# VIEWS HỖ TRỢ - GỬI TIN NHẮN
# ===========================================================================

@login_required
def support_dashboard(request):
    """Trang chính hỗ trợ"""
    user = request.user
    
    # Sửa dòng này:
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    # Lấy danh sách giáo viên và đề thi (cho học sinh)
    teacher_quizzes = None
    teachers = None
    
    if user.role == 'STUDENT':
        # Lấy đề thi mà học sinh có thể liên hệ giáo viên
        teacher_quizzes = Quiz.objects.filter(
            models.Q(is_public=True) | 
            models.Q(allowed_students=user)
        ).distinct().order_by('-created_at')[:10]
        
        # Lấy danh sách giáo viên - SỬA DÒNG NÀY
        teachers = User.objects.filter(role='TEACHER').order_by('username')[:10]
    
    # Lấy yêu cầu hỗ trợ đã gửi
    sent_tickets = SupportTicket.objects.filter(user=user).order_by('-created_at')[:5]
    
    context = {
        'user': user,
        'role': user.role,
        'teacher_quizzes': teacher_quizzes,
        'teachers': teachers,
        'sent_tickets': sent_tickets,
    }
    
    return render(request, 'support/support_dashboard.html', context)

@login_required
def contact_teacher(request, teacher_id=None, quiz_id=None):
    """Liên hệ với giáo viên"""
    user = request.user
    
    # Chỉ học sinh mới được liên hệ giáo viên
    if user.role != 'STUDENT':
        messages.error(request, "Chỉ học sinh mới có thể liên hệ giáo viên")
        return redirect('quiz:support_dashboard')
    
    teacher = None
    quiz = None
    
    # Lấy thông tin giáo viên và đề thi
    if teacher_id:
        teacher = get_object_or_404(User, id=teacher_id, role='TEACHER')
    
    if quiz_id:
        quiz = get_object_or_404(Quiz, id=quiz_id)
        teacher = quiz.created_by  # Giáo viên tạo đề thi
    
    if request.method == 'POST':
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()
        ticket_type = request.POST.get('ticket_type', 'QUESTION')
        
        if not subject or not message:
            messages.error(request, "Vui lòng điền đầy đủ thông tin!")
            return render(request, 'support/contact_teacher.html', {
                'teacher': teacher,
                'quiz': quiz,
                'ticket_types': SupportTicket.TicketType.choices,
            })
        
        try:
            # Tạo ticket hỗ trợ
            ticket = SupportTicket.objects.create(
                user=user,
                teacher=teacher,
                quiz=quiz,
                ticket_type=ticket_type,
                subject=subject,
                message=message,
                status='OPEN'
            )
            
            messages.success(request, "Đã gửi yêu cầu hỗ trợ đến giáo viên!")
            return redirect('quiz:support_ticket_detail', ticket_id=ticket.id)
            
        except Exception as e:
            messages.error(request, f"Có lỗi khi gửi yêu cầu: {str(e)}")
    
    context = {
        'user': user,
        'teacher': teacher,
        'quiz': quiz,
        'ticket_types': SupportTicket.TicketType.choices,
    }
    return render(request, 'support/contact_teacher.html', context)

@login_required
def contact_admin(request):
    """Liên hệ với quản trị viên"""
    user = request.user
    
    # Tính số ticket đang mở của user
    open_tickets_count = SupportTicket.objects.filter(
        user=user,
        status='OPEN'
    ).count()
    
    # Tính ID ticket tiếp theo
    last_ticket = SupportTicket.objects.order_by('-id').first()
    next_ticket_id = (last_ticket.id + 1) if last_ticket else 1
    
    if request.method == 'POST':
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()
        ticket_type = request.POST.get('ticket_type', 'OTHER')
        
        if not subject or not message:
            messages.error(request, "Vui lòng điền đầy đủ thông tin!")
            return redirect('quiz:contact_admin')
        
        try:
            # Tạo ticket hỗ trợ (không có teacher)
            ticket = SupportTicket.objects.create(
                user=user,
                ticket_type=ticket_type,
                subject=subject,
                message=message,
                status='OPEN'
            )
            
            messages.success(request, f"Đã tạo ticket #{ticket.id:06d}! Quản trị viên sẽ liên hệ với bạn sớm.")
            return redirect('quiz:support_ticket_detail', ticket_id=ticket.id)
            
        except Exception as e:
            messages.error(request, f"Có lỗi khi gửi yêu cầu: {str(e)}")
    
    context = {
        'user': user,
        'ticket_types': SupportTicket.TicketType.choices,
        'open_tickets_count': open_tickets_count,
        'next_ticket_id': f"{next_ticket_id:06d}",
    }
    return render(request, 'support/contact_admin.html', context)

@login_required
def support_ticket_detail(request, ticket_id):
    """Xem chi tiết ticket hỗ trợ"""
    ticket = get_object_or_404(SupportTicket, id=ticket_id)
    user = request.user
    
    # Kiểm tra quyền truy cập
    if ticket.user != user and user.role not in ['TEACHER', 'ADMIN']:
        raise PermissionDenied
    
    # Kiểm tra nếu là giáo viên, chỉ được xem ticket gửi cho mình
    if user.role == 'TEACHER' and ticket.teacher != user:
        raise PermissionDenied
    
    # Đánh dấu là đã đọc nếu là giáo viên/admin
    if user.role in ['TEACHER', 'ADMIN'] and not ticket.is_read:
        ticket.is_read = True
        ticket.save()
    
    if request.method == 'POST':
        # Giáo viên/Admin phản hồi
        if user.role in ['TEACHER', 'ADMIN']:
            response = request.POST.get('response', '').strip()
            status = request.POST.get('status', ticket.status)
            
            if response:
                ticket.admin_response = response
                ticket.admin = user if user.role == 'ADMIN' else None
                ticket.replied_by = user
                ticket.replied_at = timezone.now()
                ticket.status = status
                ticket.save()
                
                messages.success(request, "Đã gửi phản hồi!")
        else:
            # Học sinh cập nhật ticket
            new_message = request.POST.get('new_message', '').strip()
            if new_message:
                ticket.message += f"\n\n--- Cập nhật từ học sinh ({timezone.now().strftime('%d/%m/%Y %H:%M')}) ---\n{new_message}"
                ticket.is_read = False  # Đánh dấu là chưa đọc để giáo viên biết có cập nhật mới
                ticket.save()
                messages.success(request, "Đã cập nhật yêu cầu!")
    
    context = {
        'ticket': ticket,
        'can_reply': user.role in ['TEACHER', 'ADMIN'],  # Giáo viên/Admin mới được phản hồi
        'is_owner': ticket.user == user,  # Người tạo ticket
        'has_replied': ticket.replied_by is not None,  # Đã có phản hồi chưa
    }
    return render(request, 'support/ticket_detail.html', context)
# views.py - Thêm sau view support_ticket_detail
@login_required
@require_POST
def update_ticket_status(request, ticket_id):
    """View cập nhật trạng thái ticket"""
    ticket = get_object_or_404(SupportTicket, id=ticket_id)
    
    # Kiểm tra quyền: chỉ admin hoặc giáo viên được chỉ định mới được cập nhật
    user = request.user
    is_authorized = False
    
    if user.role == 'ADMIN':
        is_authorized = True
    elif user.role == 'TEACHER':
        # Giáo viên chỉ được cập nhật ticket gửi cho mình
        if ticket.teacher == user:
            is_authorized = True
    
    if not is_authorized:
        messages.error(request, "Bạn không có quyền cập nhật trạng thái ticket này.")
        return redirect('quiz:support_ticket_detail', ticket_id=ticket_id)
    
    # Lấy trạng thái mới từ form
    new_status = request.POST.get('status')
    
    if new_status and new_status in dict(SupportTicket.STATUS_CHOICES):
        old_status_display = ticket.get_status_display()
        ticket.status = new_status
        
        # Ghi lại lịch sử cập nhật
        ticket.status_history = f"{ticket.status_history or ''}\n{timezone.now().strftime('%d/%m/%Y %H:%M')} - {user.username} cập nhật từ {old_status_display} thành {ticket.get_status_display()}"
        
        # Nếu là admin và có comment, thêm vào admin_response
        admin_comment = request.POST.get('admin_comment', '').strip()
        if admin_comment and user.role == 'ADMIN':
            ticket.admin_response = f"{ticket.admin_response or ''}\n\n--- Cập nhật trạng thái ({timezone.now().strftime('%d/%m/%Y %H:%M')}) ---\n{admin_comment}"
            ticket.replied_by = user
            ticket.replied_at = timezone.now()
            ticket.is_read = False  # Đánh dấu để người dùng biết có cập nhật
        
        ticket.save()
        messages.success(request, f"Đã cập nhật trạng thái thành {ticket.get_status_display()}.")
    else:
        messages.error(request, "Trạng thái không hợp lệ.")
    
    return redirect('quiz:support_ticket_detail', ticket_id=ticket_id)
@login_required
def my_support_tickets(request):
    """Danh sách yêu cầu hỗ trợ của tôi"""
    user = request.user
    tickets = SupportTicket.objects.filter(user=user).order_by('-created_at')
    
    # Thêm filter theo status
    status_filter = request.GET.get('status')
    if status_filter:
        tickets = tickets.filter(status=status_filter)
    
    # Thêm filter theo loại ticket
    ticket_type_filter = request.GET.get('ticket_type')
    if ticket_type_filter:
        tickets = tickets.filter(ticket_type=ticket_type_filter)
    
    # Thêm search
    search_query = request.GET.get('search')
    if search_query:
        tickets = tickets.filter(
            Q(subject__icontains=search_query) | 
            Q(message__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(tickets, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'tickets': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'ticket_types': SupportTicket.TicketType.choices,
    }
    return render(request, 'support/my_ticket.html', context)
@login_required
@teacher_required
def teacher_support_inbox(request):
    """Hộp thư hỗ trợ của giáo viên"""
    teacher = request.user
    
    # Lấy các ticket gửi cho giáo viên này
    tickets = SupportTicket.objects.filter(teacher=teacher).order_by('-created_at')
    
    # Thêm filter
    status_filter = request.GET.get('status')
    if status_filter:
        tickets = tickets.filter(status=status_filter)
    
    # Tìm kiếm
    search_query = request.GET.get('search')
    if search_query:
        tickets = tickets.filter(
            Q(subject__icontains=search_query) | 
            Q(message__icontains=search_query) |
            Q(user__username__icontains=search_query)
        )
    
    # Thống kê
    total_tickets = tickets.count()
    open_tickets = tickets.filter(status='OPEN').count()
    in_progress_tickets = tickets.filter(status='IN_PROGRESS').count()
    resolved_tickets = tickets.filter(status='RESOLVED').count()
    unread_tickets = tickets.filter(is_read=False).count()
    
    # Phân loại theo loại ticket
    ticket_type_stats = {}
    for ticket_type, display_name in SupportTicket.TicketType.choices:
        count = tickets.filter(ticket_type=ticket_type).count()
        if count > 0:
            ticket_type_stats[display_name] = count
    
    # Pagination
    paginator = Paginator(tickets, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'tickets': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
        'in_progress_tickets': in_progress_tickets,
        'resolved_tickets': resolved_tickets,
        'unread_tickets': unread_tickets,
        'ticket_types': SupportTicket.TicketType.choices,
        'ticket_type_stats': ticket_type_stats,
        'teacher': teacher,
    }
    return render(request, 'support/teacher_inbox.html', context)
@login_required
@admin_required
def admin_support_dashboard(request):
    """Dashboard hỗ trợ của admin"""
    user = request.user
    
    # Lấy tất cả ticket
    tickets = SupportTicket.objects.all().order_by('-created_at')
    
    # Thêm filter
    status_filter = request.GET.get('status')
    if status_filter:
        tickets = tickets.filter(status=status_filter)
    
    # Tìm kiếm
    search_query = request.GET.get('search')
    if search_query:
        tickets = tickets.filter(
            Q(subject__icontains=search_query) | 
            Q(message__icontains=search_query) |
            Q(user__username__icontains=search_query)
        )
    
    # Thống kê
    total_tickets = tickets.count()
    open_tickets = tickets.filter(status='OPEN').count()
    in_progress_tickets = tickets.filter(status='IN_PROGRESS').count()
    resolved_tickets = tickets.filter(status='RESOLVED').count()
    closed_tickets = tickets.filter(status='CLOSED').count()
    
    # Phân loại theo loại yêu cầu
    ticket_types_stats = {}
    for ticket_type, display_name in SupportTicket.TicketType.choices:
        count = tickets.filter(ticket_type=ticket_type).count()
        if count > 0:
            ticket_types_stats[display_name] = count
    
    # Phân loại theo giáo viên (nếu có)
    teacher_tickets = {}
    teacher_tickets_data = tickets.filter(teacher__isnull=False).values('teacher__username').annotate(count=Count('id'))
    for item in teacher_tickets_data:
        teacher_tickets[item['teacher__username']] = item['count']
    
    # Phân loại theo học sinh
    student_tickets = {}
    student_tickets_data = tickets.values('user__username').annotate(count=Count('id')).order_by('-count')[:10]
    for item in student_tickets_data:
        student_tickets[item['user__username']] = item['count']
    
    # Pagination
    paginator = Paginator(tickets, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'tickets': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
        'in_progress_tickets': in_progress_tickets,
        'resolved_tickets': resolved_tickets,
        'closed_tickets': closed_tickets,
        'ticket_types_stats': ticket_types_stats,
        'teacher_tickets': teacher_tickets,
        'student_tickets': student_tickets,
        'user': user,
    }
    return render(request, 'support/admin_dashboard.html', context)
# ===========================================================================
# VIEWS KHÁM PHÁ ĐỀ THI
# ===========================================================================

@login_required
def explore(request):
    """Trang khám phá đề thi"""
    # 1. Lấy tất cả môn học
    subjects = Subject.objects.all()
    
    # 2. Lấy các đề thi CÔNG KHAI và ĐANG MỞ
    now = timezone.now()
    public_quizzes = Quiz.objects.filter(
        is_public=True,
        start_time__lte=now,
        end_time__gte=now
    ).order_by('-created_at')

    # 3. Xử lý tìm kiếm (nếu người dùng gõ vào ô search)
    query = request.GET.get('q')
    if query:
        public_quizzes = public_quizzes.filter(title__icontains=query)

    context = {
        'subjects': subjects,
        'quizzes': public_quizzes,
        'query': query
    }
    return render(request, 'pages/explore.html', context)